"""Blueprint de ativos: listagem, criação, visualização, edição e baixa."""
from datetime import datetime
from decimal import Decimal, InvalidOperation

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user, login_required
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError

from app.extensions import db
from app.models.asset import Asset, AssetType
from app.models.employee import Employee
from app.models.enums import AssetCondition, AssetStatus, UserRole
from app.models.location import Location
from app.models.maintenance import MaintenanceRecord
from app.models.movement import AssetMovement
from app.models.supplier import Supplier
from app.utils.decorators import can_manage_slug, module_required
from app.utils.spec_config import SPEC_CONFIG

assets_bp = Blueprint("assets", __name__, url_prefix="/assets")

# Ativos fazem parte do módulo "Inventário de Máquinas". Acesso: ver exige
# can_view; criar/editar/baixar/atribuir exige "Gerenciar" (can_manage).
MODULE_SLUG = "inventario-maquinas"


# ---------------------------------------------------------------------------
# Helpers de coerção
# ---------------------------------------------------------------------------
def _str_or_none(value):
    value = (value or "").strip()
    return value or None


def _int_or_none(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_decimal(value):
    value = (value or "").strip()
    if not value:
        return None
    # Aceita formato BR (1.234,56) e internacional (1234.56).
    if "," in value:
        value = value.replace(".", "").replace(",", ".")
    try:
        return Decimal(value)
    except InvalidOperation:
        return None


def _enum_or_default(enum_cls, name, default):
    try:
        return enum_cls[name]
    except (KeyError, TypeError):
        return default


def _coerce_field(field, form, name):
    t = field["type"]
    if t == "checkbox":
        return name in form
    raw = (form.get(name) or "").strip()
    if raw == "":
        return None
    if t == "number":
        try:
            return int(raw)
        except ValueError:
            return None
    if t == "select":
        try:
            return field["enum"][raw]
        except (KeyError, TypeError):
            return None
    return raw


def _sync_spec(asset, asset_type, form):
    """Cria/atualiza a spec correspondente ao tipo do ativo."""
    cfg = SPEC_CONFIG.get(asset_type.slug)
    if not cfg:
        return
    spec = getattr(asset, cfg["rel"])
    if spec is None:
        spec = cfg["model"]()
        setattr(asset, cfg["rel"], spec)
    for field in cfg["fields"]:
        name = f"spec__{asset_type.slug}__{field['key']}"
        setattr(spec, field["key"], _coerce_field(field, form, name))


def _apply_form(asset, form):
    """Popula o asset a partir do formulário. Retorna msg de erro ou None."""
    tag = (form.get("asset_tag") or "").strip()
    if not tag:
        return "Informe o patrimônio (asset_tag)."

    type_id = _int_or_none(form.get("asset_type_id"))
    asset_type = db.session.get(AssetType, type_id) if type_id else None
    if asset_type is None:
        return "Selecione um tipo de ativo válido."

    asset.asset_tag = tag
    asset.serial_number = _str_or_none(form.get("serial_number"))
    asset.asset_type_id = asset_type.id
    asset.brand = _str_or_none(form.get("brand"))
    asset.model = _str_or_none(form.get("model"))
    asset.status = _enum_or_default(AssetStatus, form.get("status"), AssetStatus.ACTIVE)
    asset.condition = _enum_or_default(
        AssetCondition, form.get("condition"), AssetCondition.GOOD
    )
    asset.purchase_date = _parse_date(form.get("purchase_date"))
    asset.warranty_expiry_date = _parse_date(form.get("warranty_expiry_date"))
    asset.purchase_price = _parse_decimal(form.get("purchase_price"))
    asset.supplier_id = _int_or_none(form.get("supplier_id"))
    asset.invoice_number = _str_or_none(form.get("invoice_number"))
    asset.location_id = _int_or_none(form.get("location_id"))
    asset.assigned_to_id = _int_or_none(form.get("assigned_to_id"))
    asset.image_url = _str_or_none(form.get("image_url"))
    asset.notes = _str_or_none(form.get("notes"))

    if asset.created_by_id is None and current_user.is_authenticated:
        asset.created_by_id = current_user.id

    _sync_spec(asset, asset_type, form)
    return None


def _form_context():
    """Dados compartilhados pelos selects do formulário."""
    return {
        "asset_types": AssetType.query.order_by(AssetType.name).all(),
        "locations": Location.query.order_by(Location.name).all(),
        "employees": Employee.query.filter_by(is_active=True)
        .order_by(Employee.name)
        .all(),
        "suppliers": Supplier.query.filter_by(is_active=True)
        .order_by(Supplier.name)
        .all(),
        "spec_config": SPEC_CONFIG,
    }


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------
def _filtered_query(args):
    """Aplica os filtros da listagem (status, tipo, busca) e devolve
    (query, filtros)."""
    query = Asset.query.filter(Asset.is_active.is_(True))
    status = args.get("status") or ""
    type_id = _int_or_none(args.get("type"))
    search = (args.get("q") or "").strip()

    if status:
        st = _enum_or_default(AssetStatus, status, None)
        if st is not None:
            query = query.filter(Asset.status == st)
    if type_id:
        query = query.filter(Asset.asset_type_id == type_id)
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Asset.asset_tag.ilike(like),
                Asset.serial_number.ilike(like),
                Asset.brand.ilike(like),
                Asset.model.ilike(like),
            )
        )
    return query.order_by(Asset.created_at.desc()), {
        "status": status, "type": type_id, "q": search,
    }


def _rows_per_page():
    try:
        return int((current_user.get_pref("appearance", {}) or {}).get("rows_per_page", 30))
    except Exception:  # noqa: BLE001
        return 30


@assets_bp.route("/")
@module_required(MODULE_SLUG)
def list_assets():
    query, filters = _filtered_query(request.args)
    page = request.args.get("page", 1, type=int)
    pagination = query.paginate(page=page, per_page=_rows_per_page(), error_out=False)
    return render_template(
        "assets/list.html",
        assets=pagination.items,
        pagination=pagination,
        asset_types=AssetType.query.order_by(AssetType.name).all(),
        filters=filters,
        can_edit=can_manage_slug(MODULE_SLUG),
    )


@assets_bp.route("/export")
@module_required(MODULE_SLUG)
def export_assets():
    """Exporta a listagem filtrada de ativos para um arquivo .xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        current_app.logger.warning("Exportação Excel indisponível: 'openpyxl' não instalado.")
        flash("Não foi possível exportar para Excel agora. Contate o administrador do sistema.", "warning")
        return redirect(url_for("assets.list_assets", **request.args))

    query, _ = _filtered_query(request.args)
    assets = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"
    headers = [
        "Patrimônio", "Série", "Tipo", "Marca", "Modelo", "Status", "Condição",
        "Responsável", "Localização", "Fornecedor", "Data compra",
        "Fim garantia", "Valor (R$)", "Nota fiscal",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    from app.utils.template_helpers import STATUS_LABELS, CONDITION_LABELS
    for a in assets:
        ws.append([
            a.asset_tag,
            a.serial_number or "",
            a.asset_type.name if a.asset_type else "",
            a.brand or "",
            a.model or "",
            STATUS_LABELS.get(a.status.value, a.status.value),
            CONDITION_LABELS.get(a.condition.value, a.condition.value),
            a.assigned_to.name if a.assigned_to else "",
            a.location.name if a.location else "",
            a.supplier.name if a.supplier else "",
            a.purchase_date.strftime("%d/%m/%Y") if a.purchase_date else "",
            a.warranty_expiry_date.strftime("%d/%m/%Y") if a.warranty_expiry_date else "",
            float(a.purchase_price) if a.purchase_price is not None else "",
            a.invoice_number or "",
        ])

    # Largura automática simples.
    for i, _h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 18

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"ativos_{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@assets_bp.route("/<int:asset_id>/qrcode.png")
@module_required(MODULE_SLUG)
def qrcode_png(asset_id):
    """PNG do QR Code que aponta para a ficha do ativo."""
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)
    try:
        import io

        import qrcode
        import qrcode.image.svg
        target = url_for("assets.view_asset", asset_id=asset.id, _external=True)
        img = qrcode.make(target, image_factory=qrcode.image.svg.SvgImage)
        buf = io.BytesIO()
        img.save(buf)
        buf.seek(0)
    except Exception:  # noqa: BLE001 — ImportError se 'qrcode' faltar
        current_app.logger.exception(
            "Falha ao gerar QR Code (verifique 'qrcode' e REINICIE o servidor)"
        )
        abort(503)
    from flask import send_file
    return send_file(buf, mimetype="image/svg+xml")


@assets_bp.route("/<int:asset_id>/label")
@module_required(MODULE_SLUG)
def label(asset_id):
    """Página de etiqueta imprimível com QR Code."""
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)
    return render_template("assets/label.html", asset=asset)


@assets_bp.route("/<int:asset_id>/termo")
@module_required(MODULE_SLUG)
def termo(asset_id):
    """Gera o termo de responsabilidade (PDF) do ativo atribuído."""
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)
    if asset.assigned_to is None:
        flash("Este ativo não tem responsável — atribua antes de gerar o termo.", "warning")
        return redirect(url_for("assets.view_asset", asset_id=asset.id))

    try:
        pdf_buf = _build_termo_pdf(asset)
    except ImportError:
        current_app.logger.warning(
            "Geração de PDF indisponível: 'reportlab' não encontrado. "
            "Instale as dependências e REINICIE o servidor."
        )
        flash("Não foi possível gerar o PDF agora. Contate o administrador do sistema.", "warning")
        return redirect(url_for("assets.view_asset", asset_id=asset.id))
    except Exception:  # noqa: BLE001 — qualquer outra falha vira aviso amigável
        current_app.logger.exception("Falha inesperada ao gerar o termo em PDF")
        flash("Não foi possível gerar o PDF agora. Contate o administrador do sistema.", "warning")
        return redirect(url_for("assets.view_asset", asset_id=asset.id))

    from flask import send_file
    return send_file(
        pdf_buf,
        as_attachment=False,
        download_name=f"termo_{asset.asset_tag}.pdf",
        mimetype="application/pdf",
    )


def _lat1(text):
    """Sanitiza texto para as fontes core do PDF (latin-1), evitando erro
    com caracteres fora do conjunto (ex.: emojis, travessão)."""
    return (text or "").encode("latin-1", "replace").decode("latin-1")


def _build_termo_pdf(asset):
    """Monta o PDF do termo de responsabilidade (fpdf2) e retorna um BytesIO."""
    import io

    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    from app.utils.settings import get_setting

    company = get_setting("company_name", "Inventário de TI")
    emp = asset.assigned_to
    spec_bits = []
    if asset.brand:
        spec_bits.append(asset.brand)
    if asset.model:
        spec_bits.append(asset.model)
    descr = " ".join(spec_bits) or (asset.asset_type.name if asset.asset_type else "Equipamento")

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(25, 20, 25)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _lat1("TERMO DE RESPONSABILIDADE"), align="C",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 7, _lat1(company), align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    pdf.set_font("Helvetica", size=11)
    texto = (
        f"Eu, {emp.name}"
        + (f" (matrícula {emp.employee_id})" if getattr(emp, "employee_id", None) else "")
        + (f", do departamento {emp.department.name}" if getattr(emp, "department", None) else "")
        + ", declaro ter recebido o equipamento abaixo descrito, comprometendo-me a "
        "zelar por sua guarda e conservação, utilizá-lo exclusivamente para fins de "
        "trabalho e devolvê-lo quando solicitado ou ao término do vínculo."
    )
    pdf.multi_cell(0, 6, _lat1(texto))
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Dados do equipamento", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=11)
    linhas = [
        f"Patrimônio: {asset.asset_tag}",
        f"Descrição: {descr}",
        f"Número de série: {asset.serial_number or '-'}",
        f"Tipo: {asset.asset_type.name if asset.asset_type else '-'}",
        f"Localização: {asset.location.name if asset.location else '-'}",
    ]
    for ln in linhas:
        pdf.cell(0, 6, _lat1(ln), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(6)
    pdf.cell(0, 6, _lat1(f"Data: {datetime.now().strftime('%d/%m/%Y')}"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(20)
    y = pdf.get_y()
    left = pdf.l_margin
    pdf.line(left, y, left + 75, y)
    pdf.line(left + 90, y, left + 160, y)
    pdf.set_y(y + 2)
    pdf.set_font("Helvetica", size=9)
    pdf.cell(90, 5, _lat1(emp.name))
    pdf.cell(0, 5, _lat1("Responsável pela entrega (TI)"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return io.BytesIO(bytes(pdf.output()))


@assets_bp.route("/<int:asset_id>")
@module_required(MODULE_SLUG)
def view_asset(asset_id):
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)
    cfg = SPEC_CONFIG.get(asset.asset_type.slug) if asset.asset_type else None
    spec = getattr(asset, cfg["rel"]) if cfg else None
    movements = (
        AssetMovement.query.filter_by(asset_id=asset.id)
        .order_by(AssetMovement.moved_at.desc())
        .all()
    )
    maintenances = (
        MaintenanceRecord.query.filter_by(asset_id=asset.id)
        .order_by(MaintenanceRecord.id.desc())
        .all()
    )
    return render_template(
        "assets/detail.html",
        asset=asset,
        spec_cfg=cfg,
        spec=spec,
        movements=movements,
        maintenances=maintenances,
        can_edit=can_manage_slug(MODULE_SLUG),
    )


@assets_bp.route("/new", methods=["GET", "POST"])
@module_required(MODULE_SLUG, manage=True)
def new_asset():
    asset = Asset()
    if request.method == "POST":
        error = _apply_form(asset, request.form)
        if error is None:
            try:
                db.session.add(asset)
                db.session.commit()
                flash("Ativo criado com sucesso.", "success")
                return redirect(url_for("assets.view_asset", asset_id=asset.id))
            except IntegrityError:
                db.session.rollback()
                error = "Já existe um ativo com esse patrimônio (asset_tag)."
        flash(error, "danger")
    return render_template(
        "assets/form.html", asset=asset, is_new=True, **_form_context()
    )


@assets_bp.route("/<int:asset_id>/edit", methods=["GET", "POST"])
@module_required(MODULE_SLUG, manage=True)
def edit_asset(asset_id):
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)
    if request.method == "POST":
        error = _apply_form(asset, request.form)
        if error is None:
            try:
                db.session.commit()
                flash("Ativo atualizado com sucesso.", "success")
                return redirect(url_for("assets.view_asset", asset_id=asset.id))
            except IntegrityError:
                db.session.rollback()
                error = "Já existe um ativo com esse patrimônio (asset_tag)."
        flash(error, "danger")
    return render_template(
        "assets/form.html", asset=asset, is_new=False, **_form_context()
    )


@assets_bp.route("/<int:asset_id>/delete", methods=["POST"])
@module_required(MODULE_SLUG, manage=True)
def delete_asset(asset_id):
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)
    # Soft delete: nunca remove fisicamente.
    asset.is_active = False
    db.session.commit()
    flash(f"Ativo {asset.asset_tag} baixado (removido logicamente).", "info")
    return redirect(url_for("assets.list_assets"))


@assets_bp.route("/<int:asset_id>/assign", methods=["GET", "POST"])
@module_required(MODULE_SLUG, manage=True)
def assign_asset(asset_id):
    """Atribui/transfere um ativo: grava histórico e atualiza o estado atual."""
    asset = db.session.get(Asset, asset_id)
    if asset is None:
        abort(404)

    if request.method == "POST":
        to_employee_id = _int_or_none(request.form.get("to_employee_id"))
        to_location_id = _int_or_none(request.form.get("to_location_id"))
        reason = _str_or_none(request.form.get("reason"))
        notes = _str_or_none(request.form.get("notes"))

        # Registra a movimentação (histórico / cadeia de custódia).
        movement = AssetMovement(
            asset_id=asset.id,
            from_employee_id=asset.assigned_to_id,
            to_employee_id=to_employee_id,
            from_location_id=asset.location_id,
            to_location_id=to_location_id,
            moved_by_id=current_user.id,
            reason=reason,
            notes=notes,
        )
        db.session.add(movement)

        # Atualiza o estado atual do ativo.
        asset.assigned_to_id = to_employee_id
        if to_location_id is not None:
            asset.location_id = to_location_id

        db.session.commit()
        flash("Movimentação registrada com sucesso.", "success")
        return redirect(url_for("assets.view_asset", asset_id=asset.id))

    return render_template(
        "assets/assign.html",
        asset=asset,
        locations=Location.query.order_by(Location.name).all(),
        employees=Employee.query.filter_by(is_active=True)
        .order_by(Employee.name)
        .all(),
    )
