"""Importação idempotente de certificados a partir do JSON do crt.sh.

O crt.sh (Certificate Transparency) devolve MUITAS entradas quase iguais:
o mesmo certificado aparece como *precert* e *leaf* (mesmo serial_number) e
cada renovação gera novas entradas. A identidade real do certificado é o
``serial_number`` — por isso o upsert é feito por ele: reimportar o mesmo
JSON não duplica, apenas atualiza.

Cada entrada traz ``name_value`` com vários domínios (SANs) separados por
``\n``; eles são normalizados na tabela certificate_domains.

Uso:
    from app.utils.cert_import import import_certificates, import_from_json_text
    resumo = import_certificates(lista_de_dicts)
    # -> {"total":.., "created":.., "updated":.., "skipped":.., "domains":..}
"""
import json
from datetime import datetime

from app.extensions import db
from app.models.certificate import Certificate, CertificateDomain


def _parse_dt(value):
    """Converte string ISO do crt.sh em datetime. Tolerante a variações
    ('2026-07-06T16:07:27' e '2026-07-06T17:05:58.227')."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip().replace("Z", "")
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                    # Formatos brasileiros (planilhas Excel manuais).
                    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                    "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _split_domains(name_value):
    """Quebra name_value em domínios únicos, normalizados (minúsculo, sem
    duplicar, preservando a ordem de aparição)."""
    if not name_value:
        return []
    seen, out = set(), []
    for raw in str(name_value).replace("\r", "\n").split("\n"):
        d = raw.strip().lower()
        if not d or d in seen:
            continue
        seen.add(d)
        out.append(d)
    return out


def _detect_env(domain):
    """Infere o ambiente a partir do nome do domínio (best-effort)."""
    d = domain.lower()
    # Verifica "rótulos" isolados por . ou - para não pegar substrings soltas.
    labels = d.replace("-", ".").split(".")
    if any(x in labels for x in ("stg", "staging", "stage")):
        return "stg"
    if any(x in labels for x in ("hml", "homolog", "homologacao", "uat", "qa")):
        return "hml"
    if any(x in labels for x in ("dev", "develop", "sandbox", "sbx")):
        return "dev"
    if any(x in labels for x in ("prd", "prod", "producao", "production")):
        return "prd"
    # Sem marcador explícito: assume produção (domínio "limpo").
    return "prd"


def _apply_domains(cert, name_value):
    """Reconstrói os domínios do certificado a partir de name_value."""
    cert.domains.clear()  # delete-orphan remove os antigos
    for dom in _split_domains(name_value):
        cert.domains.append(CertificateDomain(
            domain=dom,
            is_wildcard=dom.startswith("*."),
            environment=_detect_env(dom),
        ))
    return len(cert.domains)


def import_certificates(entries, *, commit=True):
    """Faz o upsert de uma lista de entradas do crt.sh (dicts).

    Deduplica por serial_number (precert/leaf e renovações colapsam numa
    linha). Retorna um resumo com contagens.
    """
    summary = {"total": 0, "created": 0, "updated": 0, "skipped": 0, "domains": 0}
    if not entries:
        return summary

    for entry in entries:
        summary["total"] += 1
        if not isinstance(entry, dict):
            summary["skipped"] += 1
            continue
        serial = (entry.get("serial_number") or "").strip()
        if not serial:
            summary["skipped"] += 1
            continue

        cert = Certificate.query.filter_by(serial_number=serial).first()
        is_new = cert is None
        if is_new:
            cert = Certificate(serial_number=serial, is_active=True)
            db.session.add(cert)

        cert.crtsh_id = entry.get("id")
        cert.issuer_ca_id = entry.get("issuer_ca_id")
        cert.issuer_name = entry.get("issuer_name")
        cert.common_name = entry.get("common_name")
        cert.not_before = _parse_dt(entry.get("not_before"))
        cert.not_after = _parse_dt(entry.get("not_after"))
        cert.entry_timestamp = _parse_dt(entry.get("entry_timestamp"))
        cert.result_count = entry.get("result_count")
        # Observações: só sobrescreve se vier preenchido (não apaga notas
        # manuais numa reimportação do JSON, que não traz esse campo).
        if entry.get("notes"):
            cert.notes = entry["notes"]

        summary["domains"] += _apply_domains(cert, entry.get("name_value"))
        summary["created" if is_new else "updated"] += 1

    if commit:
        db.session.commit()
    return summary


def import_from_json_text(text):
    """Importa a partir de um texto JSON (lista de objetos). Levanta
    ValueError com mensagem amigável se o JSON for inválido."""
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON inválido: {e}") from e
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ValueError("O JSON precisa ser uma lista de certificados.")
    return import_certificates(data)


# ---------------------------------------------------------------------------
# Importação de Excel (.xlsx) — planilha manual, cabeçalhos flexíveis (PT-BR)
# ---------------------------------------------------------------------------
import re  # noqa: E402
import unicodedata  # noqa: E402

# Colunas reconhecidas -> conjunto de nomes aceitos (normalizados: minúsculo,
# sem acento, sem "(...)"). Assim a planilha pode usar rótulos amigáveis.
_HEADER_ALIASES = {
    "serial": {"serial", "serial number", "serial_number", "numero de serie",
               "num serie", "serie", "nº de serie"},
    "common_name": {"nome comum", "cn", "common name", "common_name",
                    "nome_comum", "sistema", "nome", "aplicacao", "servico",
                    "certificado"},
    "domains": {"dominios", "dominio", "domains", "domain", "name_value",
                "san", "sans", "url", "urls", "host", "hostname", "endereco",
                "enderecos", "site", "domain_name", "domain name", "dns",
                "fqdn", "base_domain", "base domain"},
    "issuer": {"emissor", "ca", "issuer", "issuer_name", "autoridade",
               "autoridade certificadora"},
    "not_before": {"emissao", "not_before", "valido de", "inicio",
                   "data de emissao", "emitido em"},
    "not_after": {"validade", "vencimento", "not_after", "valido ate",
                  "expira", "expiracao", "expiration", "data de validade",
                  "data de vencimento", "expira em", "ssl_expiration_date",
                  "ssl expiration date", "ssl expiration", "expiration date",
                  "expiry", "ssl expiry", "data expiracao"},
    "notes": {"observacoes", "observacao", "notas", "notes", "obs",
              "comentarios"},
}

# Cabeçalhos da planilha-modelo oferecida para download.
XLSX_TEMPLATE_HEADERS = [
    "Serial (opcional)", "Nome comum (CN)", "Domínios (um por linha ou ; )",
    "Emissor (CA)", "Emissão (dd/mm/aaaa)", "Validade (dd/mm/aaaa)",
    "Observações",
]
XLSX_TEMPLATE_EXAMPLE = [
    "", "warren.com.br", "warren.com.br; api.prd.warren.com.br",
    "Let's Encrypt", "06/07/2026", "04/10/2026", "certificado de exemplo",
]


def _norm_header(h):
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\(.*?\)", "", s)  # remove dicas entre parênteses
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")  # remove acentos
    return re.sub(r"\s+", " ", s).strip()


def _slug(text):
    s = _norm_header(text)
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-") or "cert"


def _map_columns(header_row):
    col = {}
    for idx, h in enumerate(header_row):
        n = _norm_header(h)
        if not n:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if n in aliases and field not in col:
                col[field] = idx
    return col


def _open_ws(file_obj):
    """Abre a primeira planilha de um .xlsx (caminho ou file-like)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Biblioteca openpyxl não instalada — rode "
                         "Instalar_Dependencias.bat.")
    import io

    if hasattr(file_obj, "read"):
        data = file_obj.read()
        src = io.BytesIO(data if isinstance(data, (bytes, bytearray))
                         else str(data).encode())
    else:
        src = file_obj  # caminho no disco
    try:
        wb = load_workbook(src, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Não consegui abrir o Excel: {e}") from e
    return wb.active


def _read_xlsx(file_obj):
    """Lê a planilha e devolve uma lista de linhas normalizadas, cada uma com
    os campos do certificado + listas 'problems' (bloqueiam) e 'warnings'
    (só avisam). NÃO grava nada — base para o import direto e para a revisão.

    Chave (serial): usa a coluna de serial se existir; senão, sintética
    **por domínio** (``XLSX-<dominio>``) — assim reimportar a planilha
    atualizada atualiza o registro no lugar, em vez de duplicar.
    """
    ws = _open_ws(file_obj)
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("Planilha vazia.")

    col = _map_columns(header)
    if "domains" not in col and "common_name" not in col:
        raise ValueError("Não encontrei coluna de domínios nem de nome comum. "
                         "Baixe a planilha-modelo para ver as colunas esperadas.")
    if "not_after" not in col:
        raise ValueError("Não encontrei a coluna de validade/vencimento "
                         "(obrigatória). Baixe a planilha-modelo.")

    def _cell(row, field):
        i = col.get(field)
        if i is None or i >= len(row):
            return None
        v = row[i]
        return v if v not in (None, "") else None

    out, seen = [], {}
    rownum = 1
    for row in rows:
        rownum += 1
        if not any(_cell(row, f) for f in
                   ("serial", "common_name", "domains", "not_after")):
            continue  # linha em branco

        cn = _cell(row, "common_name")
        domains_raw = _cell(row, "domains")
        if domains_raw is not None:
            name_value = str(domains_raw).replace(";", "\n").replace(",", "\n")
        elif cn is not None:
            name_value = str(cn)
        else:
            name_value = ""
        domains = _split_domains(name_value)

        raw_after = _cell(row, "not_after")
        not_after = _parse_dt(raw_after)
        not_before = _parse_dt(_cell(row, "not_before"))
        issuer = _cell(row, "issuer")
        notes = _cell(row, "notes")

        # Bloqueia só o que não tem identidade (sem domínio/nome). Falta de
        # validade NÃO bloqueia: entra como categoria "sem data" para ser
        # filtrada e resolvida depois.
        problems, warnings = [], []
        if not domains and cn is None:
            problems.append("Sem domínio/nome")
        if not_after is None:
            if raw_after not in (None, ""):
                warnings.append("Data de validade ilegível — entra como 'sem data'")
            else:
                warnings.append("Sem data de validade — entra como 'sem data'")

        serial = _cell(row, "serial")
        serial = str(serial).strip() if serial is not None else ""
        synthetic = not serial
        if synthetic:
            base = cn or (domains[0] if domains else "cert")
            serial = f"XLSX-{_slug(base)}"

        if serial in seen:
            warnings.append(f"Repetido na planilha (linha {seen[serial]})")
        else:
            seen[serial] = rownum

        out.append({
            "row": rownum,
            "serial_number": serial,
            "synthetic": synthetic,
            "common_name": (str(cn).strip() if cn is not None
                            else (domains[0] if domains else None)),
            "name_value": "\n".join(domains),
            "domains": domains,
            "not_before": not_before,
            "not_after": not_after,
            "not_after_raw": "" if raw_after is None else str(raw_after),
            "issuer_name": str(issuer).strip() if issuer is not None else None,
            "notes": str(notes).strip() if notes is not None else None,
            "problems": problems,
            "warnings": warnings,
        })
    return out


def _entry_from_row(r):
    """Converte uma linha de _read_xlsx no dict aceito por import_certificates."""
    return {
        "serial_number": r["serial_number"],
        "common_name": r["common_name"],
        "name_value": r["name_value"],
        "issuer_name": r["issuer_name"],
        "not_before": r["not_before"],
        "not_after": r["not_after"],
        "notes": r["notes"],
    }


def parse_xlsx(file_obj):
    """Lê a planilha para REVISÃO (sem gravar). Retorna a lista de _read_xlsx."""
    return _read_xlsx(file_obj)


def import_from_xlsx(file_obj):
    """Importa direto (CLI/.bat): grava as linhas válidas e ignora as que têm
    problemas bloqueantes. Idempotente."""
    rows = _read_xlsx(file_obj)
    entries, skipped = [], 0
    for r in rows:
        if r["problems"]:
            skipped += 1
            continue
        entries.append(_entry_from_row(r))
    summary = import_certificates(entries)
    summary["skipped"] += skipped
    summary["total"] += skipped
    return summary


def import_selected_rows(rows):
    """Importa uma lista de linhas já revisadas (dicts com os campos de
    certificado). Datas podem vir como ISO (string) ou datetime."""
    entries = [{
        "serial_number": r.get("serial_number"),
        "common_name": r.get("common_name"),
        "name_value": r.get("name_value"),
        "issuer_name": r.get("issuer_name"),
        "not_before": r.get("not_before"),
        "not_after": r.get("not_after"),
        "notes": r.get("notes"),
    } for r in rows if (r.get("serial_number"))]
    return import_certificates(entries)


def build_template_workbook():
    """Cria e retorna um openpyxl Workbook com a planilha-modelo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificados"
    ws.append(XLSX_TEMPLATE_HEADERS)
    ws.append(XLSX_TEMPLATE_EXAMPLE)
    for cell in ws[1]:  # negrito no cabeçalho
        cell.font = Font(bold=True)
    widths = [24, 26, 40, 26, 18, 18, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    return wb
